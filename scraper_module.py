# -*- coding: utf-8 -*-
"""
Created on Sun Jul 13 09:09:01 2025

@author: luisa
"""

# scraper_module.py

import pandas as pd
import yfinance as yf
from alpha_vantage.fundamentaldata import FundamentalData
from openpyxl import load_workbook

ALPHA_VANTAGE_API_KEY = "H8EEMMKDOLNA0VJF"
fd = FundamentalData(key=ALPHA_VANTAGE_API_KEY, output_format="pandas")

def validar_ticker(ticker_symbol, fuente):
    try:
        if fuente == "Alpha Vantage":
            data, _ = fd.get_income_statement_annual(ticker_symbol)
            return not data.empty
        elif fuente == "Yahoo Finance":
            empresa = yf.Ticker(ticker_symbol)
            historial = empresa.history(period="1d")
            return not historial.empty
    except:
        return False

def obtener_datos_financieros(ticker_symbol, fuente):
    ticker_symbol = ticker_symbol.strip().upper()
    
    if not validar_ticker(ticker_symbol, fuente):
        raise ValueError(f"Ticker '{ticker_symbol}' no válido o sin datos disponibles en {fuente}")

    if fuente == "Alpha Vantage":
        estado_ingresos, _ = fd.get_income_statement_annual(ticker_symbol)
        balance_general, _ = fd.get_balance_sheet_annual(ticker_symbol)
        flujo_caja, _ = fd.get_cash_flow_annual(ticker_symbol)
    
        # Eliminar columna 'date' si existe
        for df in [estado_ingresos, balance_general, flujo_caja]:
            if "date" in df.columns:
                df.drop(columns=["date"], inplace=True)

    elif fuente == "Yahoo Finance":
        empresa = yf.Ticker(ticker_symbol)
        estado_ingresos = empresa.financials.T
        balance_general = empresa.balance_sheet.T
        flujo_caja = empresa.cashflow.T

    else:
        raise ValueError(f"Fuente '{fuente}' no soportada")

    return estado_ingresos, balance_general, flujo_caja

def ajustar_columnas(archivo):
    wb = load_workbook(archivo)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2
    wb.save(archivo)

def guardar_en_excel(ticker_symbol, estado_ingresos, balance_general, flujo_caja, ruta_archivo):
    with pd.ExcelWriter(ruta_archivo, engine="openpyxl") as writer:
        estado_ingresos.to_excel(writer, sheet_name="Estado de Ingresos")
        balance_general.to_excel(writer, sheet_name="Balance General")
        flujo_caja.to_excel(writer, sheet_name="Flujo de Caja")
    
    ajustar_columnas(ruta_archivo)
